import sys
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
import subprocess
import tempfile
import webbrowser
from datetime import datetime
from PIL import Image, ImageTk
import io
import webbrowser
import tempfile
from PyPDF2 import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import shutil
import subprocess



def recurso_relativo(ruta_relativa):
    """Devuelve la ruta compatible tanto para .py como para .exe"""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, ruta_relativa)
    self.fondo = Image.open(recurso_relativo("images/fondo.jpg"))

def formatear_fecha(valor):
    return valor.strftime("%d/%m/%Y") if isinstance(valor, (datetime)) else valor

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Buscador de TPN y DA")
        self.root.geometry("1100x700")
        self.root.configure(bg="#4367BC")  # Gris claro
        #bloquear la ventana
        self.root.resizable(False, False)

        # Cargar y procesar la imagen de fondo con 50% transparencia
        fondo_path = recurso_relativo("images/fondo.jpg")  # Tu imagen de fondo
        fondo_base = Image.open(fondo_path).convert("RGBA")
        fondo_base = fondo_base.resize((1100, 700))

        # Crear una imagen blanca y combinarla con el fondo para simular transparencia
        fondo_blanco = Image.new("RGBA", fondo_base.size, (255, 255, 255, 255))
        fondo_transparente = Image.blend(fondo_blanco, fondo_base, alpha=0.7)

        # Convertir a formato compatible con Tkinter
        self.fondo_img = ImageTk.PhotoImage(fondo_transparente)  # Cargar la imagen de fondo

        # Crear Label para poner el fondo (se pone primero para estar "debajo")
        self.canvas_fondo = tk.Label(self.root, image=self.fondo_img)
        self.canvas_fondo.place(x=0, y=0, relwidth=1, relheight=1)

        # Cargar el logo y colocarlo encima del fondo
        img_path = recurso_relativo("images/logo.jpg")
        imagen = Image.open(img_path).resize((200, 200), Image.Resampling.LANCZOS)
        self.logo = ImageTk.PhotoImage(imagen)

        self.logo_label = tk.Label(self.root, image=self.logo, bg="#585555", bd=0)
        self.logo_label.grid(row=0, column=4, rowspan=6, padx=10, pady=10, sticky="ne")


        tk.Label(root, text="Buscar por nombre:", bg="#F0F0F0", fg="#003366", font=("Arial", 10, "bold")).grid(row=0, column=0, padx=10, pady=10, sticky="e")
        self.entry_buscar = tk.Entry(root, width=40, bg="#A7C7E7")
        self.entry_buscar.grid(row=0, column=1, padx=5, pady=10)
        tk.Button(root, text="Buscar", command=self.buscar_archivos, bg="#B3E5FC", fg="#003366").grid(row=0, column=2, padx=5, pady=10)

        self.lista = tk.Listbox(root, width=130, height=10, bg="#B3E5FC", fg="#003366", selectbackground="#A7C7E7")
        self.lista.grid(row=1, column=0, columnspan=4, padx=10)
        self.lista.bind("<<ListboxSelect>>", self.mostrar_datos)

        tk.Button(root, text="Ir al Registro", command=self.abrir_excel, bg="#A7C7E7", fg="#003366").grid(row=2, column=2, pady=10)
        tk.Button(root, text="Imprimir", command=self.imprimir, bg="#A7C7E7", fg="#003366").grid(row=2, column=3, pady=10)

        font_dato = ("Arial", 10)
        color_bg = "#B3E5FC"

        self.label_nombre = tk.Label(root, text="Nombre completo:", font=font_dato, bg=color_bg, fg="#003366", anchor="w", width=100)
        self.label_nombre.grid(row=3, column=0, columnspan=4, sticky="w", padx=10, pady=(15, 5))

        self.label_ci = tk.Label(root, text="C.I.:", font=font_dato, bg=color_bg, fg="#003366", anchor="w", width=100)
        self.label_ci.grid(row=4, column=0, columnspan=4, sticky="w", padx=10, pady=5)

        tk.Label(root, text="TITULO EN PROVISION NACIONAL", font=("Arial", 10, "bold"), bg=color_bg, fg="#003366").grid(row=5, column=0, columnspan=4, sticky="w", padx=10, pady=(15, 0))

        self.label_carrera_tpn = tk.Label(root, text="Carrera :", font=font_dato, bg=color_bg, anchor="w", width=100)
        self.label_carrera_tpn.grid(row=6, column=0, columnspan=4, sticky="w", padx=15)

        self.label_tpn = tk.Label(root, text="Nota :", font=font_dato, bg=color_bg, anchor="w", width=100)
        self.label_tpn.grid(row=7, column=0, columnspan=4, sticky="w", padx=15)

        self.label_fecha_tpn = tk.Label(root, text="Fecha :", font=font_dato, bg=color_bg, anchor="w", width=100)
        self.label_fecha_tpn.grid(row=8, column=0, columnspan=4, sticky="w", padx=15)

        self.label_titulo_tpn = tk.Label(root, text="Título :", font=font_dato, bg=color_bg, anchor="w", width=100)
        self.label_titulo_tpn.grid(row=9, column=0, columnspan=4, sticky="w", padx=15)

        self.label_estado_tpn = tk.Label(root, text="Certificado TPN: ", font=font_dato, bg=color_bg, anchor="w", width=100)
        self.label_estado_tpn.grid(row=10, column=0, columnspan=4, sticky="w", padx=15)
        #agregar boton subir certificado
        self.boton_subir_certificado_tpn = tk.Button(root, text="Subir Certificado", font=font_dato, bg="#003366", fg="white", command=self.subir_certificado_tpn)
        self.boton_subir_certificado_tpn.grid(row=10, column=3, sticky="e", padx=15)

        tk.Label(root, text="DIPLOMA ACADEMICO", font=("Arial", 10, "bold"), bg=color_bg, fg="#003366").grid(row=11, column=0, columnspan=4, sticky="w", padx=10, pady=(15, 0))

        self.label_carrera_da = tk.Label(root, text="Carrera :", font=font_dato, bg=color_bg, anchor="w", width=100)
        self.label_carrera_da.grid(row=12, column=0, columnspan=4, sticky="w", padx=15)

        self.label_da = tk.Label(root, text="Nota :", font=font_dato, bg=color_bg, anchor="w", width=100)
        self.label_da.grid(row=13, column=0, columnspan=4, sticky="w", padx=15)

        self.label_fecha_da = tk.Label(root, text="Fecha :", font=font_dato, bg=color_bg, anchor="w", width=100)
        self.label_fecha_da.grid(row=14, column=0, columnspan=4, sticky="w", padx=15)

        self.label_titulo_da = tk.Label(root, text="Título :", font=font_dato, bg=color_bg, anchor="w", width=100)
        self.label_titulo_da.grid(row=15, column=0, columnspan=4, sticky="w", padx=15)

        self.label_estado_da = tk.Label(root, text="Certificado DA: ", font=font_dato, bg=color_bg, anchor="w", width=100)
        self.label_estado_da.grid(row=16, column=0, columnspan=4, sticky="w", padx=15)
        #agregar boton subir certificado
        self.boton_subir_certificado_da = tk.Button(root, text="Subir Certificado", font=font_dato, bg="#003366", fg="white", command=self.subir_certificado_da)
        self.boton_subir_certificado_da.grid(row=16, column=3, sticky="e", padx=15)

        #boton para ver los certificados
        self.boton_ver_certificado_da = tk.Button(root, text="Ver Certificado", font=font_dato, bg="#003366", fg="white", command=self.ver_certificados)
        self.boton_ver_certificado_da.grid(row=16, column=4, sticky="e", padx=15)
        self.registros = []



    def buscar_archivos(self):
        criterio = self.entry_buscar.get().strip().lower()
        if not criterio:
            messagebox.showwarning("Atención", "Ingrese un nombre para buscar.")
            return

        carpeta = filedialog.askdirectory(title="Selecciona la carpeta con archivos Excel")
        if not carpeta:
            return

        self.lista.delete(0, tk.END)
        self.registros.clear()

        for archivo in os.listdir(carpeta):
            if archivo.endswith(".xlsx") or archivo.endswith(".xlsm"):
                ruta = os.path.join(carpeta, archivo)
                try:
                    wb = load_workbook(ruta, read_only=True, data_only=True)
                    for hoja in wb.worksheets:
                        for i, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
                            if fila and len(fila) >= 12 and fila[2] and criterio in str(fila[2]).lower():
                                registro = {
                                    "archivo": ruta,
                                    "hoja": hoja.title,
                                    "fila": i,
                                    "nombre": fila[2],
                                    "ci": fila[3],
                                    "carrera_tpn": fila[4],
                                    "tpn": fila[5],
                                    "fecha_tpn": fila[6],
                                    "titulo_tpn": fila[7],
                                    "carrera_da": fila[8],
                                    "da": fila[9],
                                    "fecha_da": fila[10],
                                    "titulo_da": fila[11],
                                }
                                self.registros.append(registro)
                                texto = f"{registro['nombre']} | CI: {registro['ci']} | TPN: {registro['tpn']} | DA: {registro['da']} ({os.path.basename(ruta)})"
                                self.lista.insert(tk.END, texto)
                    wb.close()
                except Exception as e:
                    print(f"Error al procesar {archivo}: {e}")

        if not self.registros:
            messagebox.showinfo("Resultado", "No se encontraron coincidencias.")

    def mostrar_datos(self, event):
        if not self.lista.curselection():
            return
        index = self.lista.curselection()[0]
        reg = self.registros[index]

        self.label_nombre.config(text=f"Nombre completo: {reg['nombre']}")
        self.label_ci.config(text=f"C.I.: {reg['ci']}")
        self.label_carrera_tpn.config(text=f"Carrera TPN: {reg['carrera_tpn']}")
        self.label_tpn.config(text=f"Nota TPN: {reg['tpn']}")
        self.label_fecha_tpn.config(text=f"Fecha TPN: {formatear_fecha(reg['fecha_tpn'])}")
        self.label_titulo_tpn.config(text=f"Título TPN: {reg['titulo_tpn']}")
        # Verificación de Título TPN
        titulo_tpn = str(reg['titulo_tpn']).strip().lower() if reg.get('titulo_tpn') else ""
        if titulo_tpn and titulo_tpn != "sin numero":
            self.label_estado_tpn.config(text="Certificado TPN: VERIFICADO", fg="green")
        else:
            self.label_estado_tpn.config(text="Certificado TPN: NO VERIFICADO", fg="red")

        self.label_carrera_da.config(text=f"Carrera DA: {reg['carrera_da']}")
        self.label_da.config(text=f"Nota DA: {reg['da']}")
        self.label_fecha_da.config(text=f"Fecha DA: {formatear_fecha(reg['fecha_da'])}")
        self.label_titulo_da.config(text=f"Título DA: {reg['titulo_da']}")
        # Verificación de Título DA
        titulo_da = str(reg['titulo_da']).strip().lower() if reg['titulo_da'] is not None else ""
        if titulo_da and titulo_da != "sin numero" :
            self.label_estado_da.config(text="Certificado DA: VERIFICADO", fg="green")
        else:
            self.label_estado_da.config(text="Certificado DA: NO VERIFICADO", fg="red")

    def abrir_excel(self):
        if not self.lista.curselection():
            messagebox.showinfo("Atención", "Seleccione un registro primero.")
            return
        reg = self.registros[self.lista.curselection()[0]]
        try:
            if sys.platform == "win32":
                os.startfile(reg["archivo"])
            else:
                subprocess.call(["open" if sys.platform == "darwin" else "xdg-open", reg["archivo"]])
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el archivo: {e}")

    def imprimir(self):
        if not self.lista.curselection():
            messagebox.showinfo("Atención", "Seleccione un registro primero.")
            return
        reg = self.registros[self.lista.curselection()[0]]

        # Crear un canvas en memoria con tamaño CARTA
        packet = io.BytesIO()
        c = canvas.Canvas(packet, pagesize=letter)

        def texto_centrado(y, texto, size=12, bold=False):
                font = "Helvetica-Bold" if bold else "Helvetica"
                c.setFont(font, size)
                text_width = c.stringWidth(texto, font, size)
                x = (letter[0] - text_width) / 2  # Centrar horizontalmente en carta (612 pts)
                c.drawString(x, y, texto)

        y = 600  # Ajuste vertical inicial para carta
        texto_centrado(y, "Informe del Registro", size=14, bold=True)
        y -= 30

        def line(label, value):
            nonlocal y
            texto = f"{label}: {value}"
            texto_centrado(y, texto)
            y -= 20

        line("Nombre completo", reg['nombre'])
        line("C.I.", reg['ci'])

        y -= 20
        texto_centrado(y, "TITULO EN PROVISION NACIONAL", bold=True)
        y -= 20

        line("Carrera", reg['carrera_tpn'])
        line("Nota", reg['tpn'])
        line("Fecha", formatear_fecha(reg['fecha_tpn']))
        line("Título", reg['titulo_tpn'])
        line("Certificado ", "VERIFICADO ✓" if reg['titulo_tpn'] and str(reg['titulo_tpn']).strip().lower() != "sin numero" else "NO VERIFICADO ✗")

        y -= 20
        texto_centrado(y, "DIPLOMA ACADEMICO", bold=True)
        y -= 20

        line("Carrera", reg['carrera_da'])
        line("Nota", reg['da'])
        line("Fecha", formatear_fecha(reg['fecha_da']))
        line("Título", reg['titulo_da'])
        line("Certificado ", "VERIFICADO ✓" if reg['titulo_da'] and str(reg['titulo_da']).strip().lower() != "sin numero" else "NO VERIFICADO ✗")

        c.save()
        packet.seek(0)

        # Cargar plantilla y superponer datos
        overlay_pdf = PdfReader(packet)
        plantilla_pdf = PdfReader(open("plantilla.pdf", "rb"))
        writer = PdfWriter()

        page = plantilla_pdf.pages[0]
        page.merge_page(overlay_pdf.pages[0])
        writer.add_page(page)

        # Guardar y abrir
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as f:
            writer.write(f)
            webbrowser.open(f.name)

####
    def subir_certificado_tpn(self):
            if not self.lista.curselection():
                messagebox.showinfo("Atención", "Seleccione un registro primero.")
                return
            reg = self.registros[self.lista.curselection()[0]]

            ruta_pdf = filedialog.askopenfilename(title="Seleccionar certificado TPN", 
                                                filetypes=[("PDF Files", "*.pdf")])
            if ruta_pdf:
                carpeta_certificados = "certificados"
                os.makedirs(carpeta_certificados, exist_ok=True)

                nombre_archivo = f"certificado_tpn_{reg['ci']}.pdf"
                destino = os.path.join(carpeta_certificados, nombre_archivo)
                try:
                    shutil.copy2(ruta_pdf, destino)
                    messagebox.showinfo("Éxito", f"Certificado TPN guardado para {reg['nombre']}")
                except Exception as e:
                    messagebox.showerror("Error", f"No se pudo guardar el certificado: {e}")

    def subir_certificado_da(self):
            if not self.lista.curselection():
                messagebox.showinfo("Atención", "Seleccione un registro primero.")
                return
            reg = self.registros[self.lista.curselection()[0]]

            ruta_pdf = filedialog.askopenfilename(title="Seleccionar certificado DA", 
                                                filetypes=[("PDF Files", "*.pdf")])
            if ruta_pdf:
                carpeta_certificados = "certificados"
                os.makedirs(carpeta_certificados, exist_ok=True)

                nombre_archivo = f"certificado_da_{reg['ci']}.pdf"
                destino = os.path.join(carpeta_certificados, nombre_archivo)
                try:
                    shutil.copy2(ruta_pdf, destino)
                    messagebox.showinfo("Éxito", f"Certificado DA guardado para {reg['nombre']}")
                except Exception as e:
                    messagebox.showerror("Error", f"No se pudo guardar el certificado: {e}")

    def ver_certificados(self):
            if not self.lista.curselection():
                messagebox.showinfo("Atención", "Seleccione un registro primero.")
                return
            reg = self.registros[self.lista.curselection()[0]]

            carpeta_certificados = "certificados"
            archivos = []
            archivo_tpn = os.path.join(carpeta_certificados, f"certificado_tpn_{reg['ci']}.pdf")
            archivo_da = os.path.join(carpeta_certificados, f"certificado_da_{reg['ci']}.pdf")

            if os.path.exists(archivo_tpn):
                archivos.append(archivo_tpn)
            if os.path.exists(archivo_da):
                archivos.append(archivo_da)

            if not archivos:
                messagebox.showinfo("Información", "No se encontraron certificados para este usuario.")
                return

            for archivo in archivos:
                try:
                    if sys.platform == "win32":
                        os.startfile(archivo)
                    elif sys.platform == "darwin":
                        subprocess.call(["open", archivo])
                    else:
                        subprocess.call(["xdg-open", archivo])
                except Exception as e:
                    messagebox.showerror("Error", f"No se pudo abrir el archivo: {e}")



if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()